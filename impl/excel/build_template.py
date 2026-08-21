#!/usr/bin/env python3
"""Build the Excel template for case el/mean-ylt-10k.

The workbook IS the implementation: the harness writes inputs into named ranges
and reads named outputs back. Formulas are laid out cell by cell rather than as
one nested expression, so a divergence can be traced to a step (spec SS 7.3).

Three strategies are computed side by side, because the interesting question is
not whether a spreadsheet can average a column but whether the ways a modeller
might do it agree (spec SS 7.3):

  EL_SUM       =SUM(range)/n     -- the value under test (spec SS 3.1)
  EL_AVERAGE   =AVERAGE(range)   -- the obvious alternative
  EL_RUNNING   running total carried down a helper column

The third is reported, not a candidate. It is the layout hand-built layer models
overwhelmingly use, so it is worth showing a reader whether it agrees.
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[2]
CSV_PATH = ROOT / "conformance" / "data" / "ylt-10k.csv"
OUT = Path(__file__).resolve().parent / "templates" / "el_mean.xlsx"

BLUE = Font(name="Arial", color="0000FF")          # input
BLACK = Font(name="Arial")                          # formula
BOLD = Font(name="Arial", bold=True)
YELLOW = PatternFill("solid", fgColor="FFFF00")     # key output
NUM = "#,##0.000000"


def build() -> Workbook:
    """Construct the template workbook in memory."""
    with CSV_PATH.open() as fh:
        losses = [r["loss"] for r in csv.DictReader(fh)]
    n = len(losses)
    last = n + 1  # data occupies rows 2..n+1

    wb = Workbook()

    # ---- YLT sheet: data plus the running-total diagnostic ---------------
    ws = wb.active
    ws.title = "YLT"
    for col, head in enumerate(["year", "loss", "S (running sum)"], start=1):
        ws.cell(row=1, column=col, value=head).font = BOLD
    ws.freeze_panes = "A2"

    for i, loss in enumerate(losses, start=2):
        ws.cell(row=i, column=1, value=i - 1).font = BLACK
        c = ws.cell(row=i, column=2, value=float(loss))
        c.font = BLUE
        c.number_format = "#,##0.00"

    ws["C2"] = "=B2"
    for r in range(3, last + 1):
        ws[f"C{r}"] = f"=C{r-1}+B{r}"
    for r in range(2, last + 1):
        ws.cell(row=r, column=3).number_format = NUM

    ws.column_dimensions["A"].width = 8
    for col in "BC":
        ws.column_dimensions[col].width = 22

    # ---- Calc sheet: named inputs and outputs ----------------------------
    cs = wb.create_sheet("Calc")
    cs["A1"] = "ILS conformance -- case el/mean-ylt-10k"
    cs["A1"].font = Font(name="Arial", bold=True, size=13)
    cs["A2"] = "Blue = harness input. Yellow = harness output. Black = formula."
    cs["A2"].font = Font(name="Arial", italic=True, size=9)

    # (label, kind, formula_or_value, note). Addresses are captured as the
    # rows are written -- hardcoding them is how the named ranges drift out of
    # step with the layout, which is a silent wrong-answer bug, not a crash.
    layout = [
        ("Inputs", "head", None, None),
        ("N_YEARS", "input", n, "Written by the harness from case.operation.n_years"),
        ("FIRST_ROW", "input", 2, "First data row on sheet YLT"),
        ("LAST_ROW", "input", last, "Last data row on sheet YLT"),
        ("Intermediates", "head", None, None),
        ("SUM_LOSSES", "calc", f"=SUM(YLT!B2:B{last})", "The spreadsheet's own accumulation"),
        ("COUNT_ROWS", "calc", f"=COUNT(YLT!B2:B{last})", "Guards an off-by-one in the range"),
        ("S_FINAL", "calc", f"=YLT!C{last}", "Running total, final row"),
        ("Outputs", "head", None, None),
        ("EL_SUM", "out", "={SUM_LOSSES}/{N_YEARS}", "Spec SS 3.1 -- the value under test"),
        ("EL_AVERAGE", "out", f"=AVERAGE(YLT!B2:B{last})", "AVERAGE(range)"),
        ("EL_RUNNING", "out", "={S_FINAL}/{N_YEARS}", "Reported: running-total column (SS 7.3)"),
    ]

    addr: dict[str, str] = {}
    r = 4
    for label, kind, value, note in layout:
        if kind == "head":
            cs.cell(row=r, column=1, value=label).font = BOLD
            r += 1
            continue
        addr[label] = f"B{r}"
        cs.cell(row=r, column=1, value=label).font = Font(name="Arial")
        # Formulas reference earlier outputs by label, resolved here, so a
        # layout change cannot leave a dangling reference behind.
        resolved = value.format(**addr) if isinstance(value, str) else value
        cell = cs.cell(row=r, column=2, value=resolved)
        cell.number_format = NUM
        cell.font = BLUE if kind == "input" else BLACK
        if kind == "out":
            cell.fill = YELLOW
        note_cell = cs.cell(row=r, column=3, value=note)
        note_cell.font = Font(name="Arial", size=9, color="666666")
        note_cell.alignment = Alignment(horizontal="left")
        r += 1

    cs.column_dimensions["A"].width = 20
    cs.column_dimensions["B"].width = 24
    cs.column_dimensions["C"].width = 52

    # ---- Named ranges: the harness contract ------------------------------
    for name, ref in addr.items():
        wb.defined_names.add(DefinedName(name, attr_text=f"Calc!${ref[0]}${ref[1:]}"))

    wb.properties.creator = "build_template.py"
    return wb


# --- staleness check --------------------------------------------------------
#
# The committed workbook must match this builder (spec SS 7.1). It is NOT
# compared byte for byte: an .xlsx is a zip of XML, and its bytes depend on the
# openpyxl version and on the platform's deflate implementation, so a byte
# comparison fails for reasons that have nothing to do with the workbook. What
# matters is the implementation -- the formulas, the values and the named ranges
# the harness binds to. Formatting is cosmetic and deliberately not compared.


def content(wb: Workbook) -> dict:
    """The parts of a workbook that make it an implementation."""
    return {
        "sheets": wb.sheetnames,
        "cells": {
            f"{ws.title}!{cell.coordinate}": cell.value
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        },
        "names": {
            name: str(wb.defined_names[name].attr_text)
            for name in sorted(wb.defined_names)
        },
    }


def differences(committed: dict, fresh: dict) -> list[str]:
    """First few concrete differences, in reader-friendly terms."""
    out: list[str] = []
    if committed["sheets"] != fresh["sheets"]:
        out.append(f"sheets: committed {committed['sheets']}, builder {fresh['sheets']}")

    for key in sorted(set(committed["names"]) | set(fresh["names"])):
        a, b = committed["names"].get(key), fresh["names"].get(key)
        if a != b:
            out.append(f"named range {key}: committed {a!r}, builder {b!r}")

    ca, cb = committed["cells"], fresh["cells"]
    for key in sorted(set(ca) | set(cb), key=lambda k: (k.split("!")[0], k)):
        a, b = ca.get(key), cb.get(key)
        if a != b:
            out.append(f"{key}: committed {a!r}, builder {b!r}")
            if len(out) > 20:
                out.append("... further differences not listed")
                return out
    return out


def check() -> int:
    if not OUT.exists():
        print(f"{OUT.relative_to(ROOT)} is missing; run without --check to build it",
              file=sys.stderr)
        return 1

    committed = content(load_workbook(OUT))
    fresh = content(build())
    diffs = differences(committed, fresh)
    if diffs:
        print(f"{OUT.relative_to(ROOT)} does not match build_template.py:",
              file=sys.stderr)
        for d in diffs:
            print(f"  {d}", file=sys.stderr)
        print("Rebuild it with `just build-excel` and commit the result.",
              file=sys.stderr)
        return 1

    print(f"{OUT.relative_to(ROOT)}: in sync with builder "
          f"({len(fresh['cells'])} cells, {len(fresh['names'])} named ranges)")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(prog="build_template")
    ap.add_argument("--check", action="store_true",
                    help="verify the committed workbook matches this builder, "
                         "without writing it (spec SS 7.1)")
    args = ap.parse_args()

    if args.check:
        return check()

    wb = build()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
