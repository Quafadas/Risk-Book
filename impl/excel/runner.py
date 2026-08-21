#!/usr/bin/env python3
"""Excel harness (spec SS 7).

Writes case inputs into named ranges, recalculates, reads named outputs back.

INDEPENDENCE (spec SS 6.2): this module must never import from impl/python. It
drives a spreadsheet; it does not compute expected loss. If it ever borrows the
reference implementation, agreement between the Excel and Python columns stops
meaning anything. CI enforces this -- see tools/check_independence.py.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[1]
TEMPLATES = HERE / "templates"

# Case family -> template. One template serves a family; the harness varies the
# inputs, not the formulas (spec SS 7.1).
BINDINGS = {"el": "el_mean.xlsx"}


def engine_version() -> str:
    """Engine identity for the envelope (spec SS 7.2).

    Reported as `libreoffice <version>`, never as `excel`: LibreOffice is a
    different implementation of Excel semantics.
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        return "libreoffice unknown"
    try:
        out = subprocess.run([soffice, "--version"], capture_output=True,
                             text=True, timeout=60).stdout.strip()
    except (OSError, subprocess.SubprocessError):
        return "libreoffice unknown"
    # "LibreOffice 24.2.7.2 420(Build:2)" -> "libreoffice 24.2.7.2"
    parts = out.split()
    return f"libreoffice {parts[1]}" if len(parts) > 1 else "libreoffice unknown"


def recalculate(path: Path, timeout: int = 300) -> None:
    """Force a full recalculation via LibreOffice headless.

    LibreOffice is a *different implementation* of Excel semantics. Results are
    reported as `libreoffice <version>`, never as `excel`; a claim about Excel
    needs a run on licensed Excel, which this harness does not do.
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice is None:
        raise RuntimeError("soffice not found; install LibreOffice or use --cached")

    with tempfile.TemporaryDirectory() as tmp:
        env = {**os.environ, "HOME": tmp}
        subprocess.run(
            [soffice, "--headless", "--norestore", f"-env:UserInstallation=file://{tmp}/li",
             "--convert-to", "xlsx", "--outdir", tmp, str(path)],
            check=True, capture_output=True, timeout=timeout, env=env,
        )
        produced = Path(tmp) / path.name
        if not produced.exists():
            raise RuntimeError(f"recalculation produced no output for {path.name}")
        shutil.copy(produced, path)


def read_named(wb, name: str):
    dn = wb.defined_names[name]
    (sheet, ref), = dn.destinations
    return wb[sheet][ref.replace("$", "")].value


def write_named(wb, name: str, value) -> None:
    dn = wb.defined_names[name]
    (sheet, ref), = dn.destinations
    wb[sheet][ref.replace("$", "")] = value


def run_case(case_path: Path, *, cached: bool = False) -> dict:
    case = json.loads(case_path.read_text(encoding="utf-8"))
    family = case["id"].split("/")[0]
    template = TEMPLATES / BINDINGS[family]
    if not template.exists():
        return {"case": case["id"], "impl": "excel", "status": "unsupported",
                "reason": f"no template bound for family '{family}'"}

    with tempfile.TemporaryDirectory() as tmp:
        work = Path(tmp) / template.name
        shutil.copy(template, work)

        if not cached:
            wb = load_workbook(work)  # formulas, no cached values
            write_named(wb, "N_YEARS", case["operation"]["n_years"])
            wb.save(work)
            recalculate(work)

        wb = load_workbook(work, data_only=True)
        outputs = {n: read_named(wb, n) for n in
                   ("EL_SUM", "EL_AVERAGE", "EL_RUNNING", "COUNT_ROWS")}

    missing = sorted(k for k, v in outputs.items() if v is None)
    if missing:
        return {"case": case["id"], "impl": "excel", "status": "error",
                "reason": f"no recalculated value for {', '.join(missing)}; "
                          "the engine did not evaluate the workbook"}

    if outputs["COUNT_ROWS"] != case["input"]["rows"]:
        return {"case": case["id"], "impl": "excel", "status": "error",
                "reason": f"template covers {outputs['COUNT_ROWS']} rows, "
                          f"case declares {case['input']['rows']}"}

    # EL_SUM is the value under test (spec SS 3.2: =SUM(range) is the
    # spreadsheet's standard summation). The other two are reported alongside it
    # so a reader can see whether the layout they use agrees (spec SS 7.3).
    value = float(outputs["EL_SUM"])
    expected = float(Decimal(case["expected"]["exact_decimal"]))

    def relative_error(v: float) -> float:
        return abs(v - expected) / abs(expected)

    err = relative_error(value)

    return {
        "case": case["id"],
        "impl": "excel",
        "engine": "cached" if cached else engine_version(),
        "spec_version": "1.0",
        "status": "pass" if err <= case["tolerance"]["rel"] else "fail",
        "value": value,
        "relative_error": err,
        "strategies": {
            k: {"value": float(v), "relative_error": relative_error(float(v))}
            for k, v in outputs.items() if k.startswith("EL_")
        },
    }


def main() -> int:
    ap = argparse.ArgumentParser(prog="ils-el-excel")
    ap.add_argument("cases", nargs="*", type=Path)
    ap.add_argument("--cached", action="store_true",
                    help="read stored values without recalculating (spec SS 7.2)")
    args = ap.parse_args()

    paths = args.cases or sorted((ROOT / "conformance" / "cases").rglob("*.json"))
    results = [run_case(p, cached=args.cached) for p in paths]
    print(json.dumps(results, indent=2))
    return 0 if all(r.get("status") in {"pass", "unsupported"} for r in results) else 1


if __name__ == "__main__":
    sys.exit(main())
